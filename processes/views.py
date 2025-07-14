"""
Views for legal processes application.
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Process
from .forms import ProcessForm
import datetime


def remove_tz(dt):
    if isinstance(dt, datetime.datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


@login_required
@permission_required('processes.view_process', raise_exception=True)
def process_list(request):
    """Display list of processes with search and pagination."""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    processes = Process.objects.all()
    
    # Apply search filter
    if search_query:
        processes = processes.filter(
            Q(process_number__icontains=search_query) |
            Q(process_class__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(judge__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        processes = processes.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(processes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Process.PROCESS_STATUS_CHOICES,
        'can_edit': request.user.has_perm('processes.change_process'),
        'can_delete': request.user.has_perm('processes.delete_process'),
    }
    
    return render(request, 'processes/process_list.html', context)


@login_required
@permission_required('processes.view_process', raise_exception=True)
def process_detail(request, pk):
    """Display process details."""
    process = get_object_or_404(Process, pk=pk)
    parties = process.parties.all()
    
    context = {
        'process': process,
        'parties': parties,
        'can_edit': request.user.has_perm('processes.change_process'),
        'can_delete': request.user.has_perm('processes.delete_process'),
    }
    
    return render(request, 'processes/process_detail.html', context)


@login_required
@permission_required('processes.add_process', raise_exception=True)
def process_create(request):
    """Create a new process."""
    if request.method == 'POST':
        form = ProcessForm(request.POST)
        if form.is_valid():
            process = form.save()
            messages.success(request, f'Process {process.process_number} created successfully.')
            return redirect('processes:process_detail', pk=process.pk)
    else:
        form = ProcessForm()
    
    context = {
        'form': form,
        'title': 'Create Process',
    }
    
    return render(request, 'processes/process_form.html', context)


@login_required
@permission_required('processes.change_process', raise_exception=True)
def process_update(request, pk):
    """Update an existing process."""
    process = get_object_or_404(Process, pk=pk)
    
    if request.method == 'POST':
        form = ProcessForm(request.POST, instance=process)
        if form.is_valid():
            process = form.save()
            messages.success(request, f'Process {process.process_number} updated successfully.')
            return redirect('processes:process_detail', pk=process.pk)
    else:
        form = ProcessForm(instance=process)
    
    context = {
        'form': form,
        'process': process,
        'title': 'Update Process',
    }
    
    return render(request, 'processes/process_form.html', context)


@login_required
@permission_required('processes.delete_process', raise_exception=True)
def process_delete(request, pk):
    """Delete a process."""
    process = get_object_or_404(Process, pk=pk)
    
    if request.method == 'POST':
        process_number = process.process_number
        process.delete()
        messages.success(request, f'Process {process_number} deleted successfully.')
        return redirect('process_list')
    
    context = {
        'process': process,
    }
    
    return render(request, 'processes/process_confirm_delete.html', context)


@login_required
@permission_required('processes.view_process', raise_exception=True)
def export_processes(request):
    """Export processes to Excel file."""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    processes = Process.objects.all()
    
    # Apply filters
    if search_query:
        processes = processes.filter(
            Q(process_number__icontains=search_query) |
            Q(process_class__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(judge__icontains=search_query)
        )
    
    if status_filter:
        processes = processes.filter(status=status_filter)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Legal Processes"
    
    # Define headers
    headers = [
        'Process Number',
        'Status',
        'Type',
        'Class',
        'Subject',
        'Judge',
        'Court',
        'Jurisdiction',
        'District',
        'Action Value',
        'Distribution Date',
        'Created At',
    ]
    
    # Style for headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row, process in enumerate(processes, 2):
        ws.cell(row=row, column=1, value=process.process_number)
        ws.cell(row=row, column=2, value=process.get_status_display())
        ws.cell(row=row, column=3, value=process.get_process_type_display())
        ws.cell(row=row, column=4, value=process.process_class)
        ws.cell(row=row, column=5, value=process.subject)
        ws.cell(row=row, column=6, value=process.judge)
        ws.cell(row=row, column=7, value=process.court)
        ws.cell(row=row, column=8, value=process.jurisdiction)
        ws.cell(row=row, column=9, value=process.district)
        ws.cell(row=row, column=10, value=float(process.action_value))
        distribution_date = remove_tz(process.distribution_date) if process.distribution_date else None
        created_at = remove_tz(process.created_at) if process.created_at else None
        ws.cell(row=row, column=11, value=distribution_date)
        ws.cell(row=row, column=12, value=created_at)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=legal_processes.xlsx'
    
    wb.save(response)
    return response
